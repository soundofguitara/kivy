# excel_manager.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime
from kivy.app import App # Pour obtenir le chemin user_data_dir
from kivy.logger import Logger # Importer Logger

EXCEL_FILE_NAME = 'warehouse_log.xlsx'
SHEET_NAME = 'InventoryLog'

def get_excel_path():
    """Retourne le chemin complet vers le fichier Excel."""
    # Utiliser user_data_dir pour un stockage approprié sur toutes les plateformes
    try:
        # Si l'application Kivy est en cours d'exécution
        app_dir = App.get_running_app().user_data_dir
        # Créer le répertoire s'il n'existe pas (plus sûr)
        os.makedirs(app_dir, exist_ok=True)
    except AttributeError:
        # Sinon (par exemple, lors de l'exécution de scripts autonomes ou de tests)
        app_dir = os.path.dirname(os.path.abspath(__file__))
        Logger.warning(f"EXCEL: Pas d'application Kivy active. Utilisation du répertoire local: {app_dir}")
        # Créer le répertoire s'il n'existe pas (également pour le cas hors-app)
        os.makedirs(app_dir, exist_ok=True)
    excel_path = os.path.join(app_dir, EXCEL_FILE_NAME)
    # Logger.info(f"EXCEL: Chemin du fichier Excel utilisé: {excel_path}") # Optionnel
    return excel_path

def init_excel():
    """Initialise le fichier Excel et crée l'en-tête si nécessaire."""
    excel_path = get_excel_path()
    Logger.info(f"EXCEL: Initialisation du fichier Excel à: {excel_path}")
    try:
        if not os.path.exists(excel_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = SHEET_NAME
            # Définir les en-têtes
            headers = [
                "ID_DB", "Action", "Timestamp", "Palette", "Produit", "Prix",
                "Date Expiration", "Lot", "Boites/Colis", "Emplacement"
            ]
            sheet.append(headers)
            # Mise en forme des en-têtes (Gras, Centré)
            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                # Ajuster la largeur de colonne (approximatif)
                column_letter = get_column_letter(col_num)
                sheet.column_dimensions[column_letter].width = 18

            workbook.save(excel_path)
            Logger.info(f"EXCEL: Fichier '{EXCEL_FILE_NAME}' créé avec succès.")
        else:
            # Vérifier si la feuille existe, sinon la créer (pour les fichiers existants corrompus)
            workbook = openpyxl.load_workbook(excel_path)
            if SHEET_NAME not in workbook.sheetnames:
                sheet = workbook.create_sheet(SHEET_NAME)
                headers = [
                    "ID_DB", "Action", "Timestamp", "Palette", "Produit", "Prix",
                    "Date Expiration", "Lot", "Boites/Colis", "Emplacement"
                ]
                sheet.append(headers)
                for col_num, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    column_letter = get_column_letter(col_num)
                    sheet.column_dimensions[column_letter].width = 18
                workbook.save(excel_path)
                Logger.info(f"EXCEL: Feuille '{SHEET_NAME}' ajoutée au fichier existant.")
            else:
                 Logger.info(f"EXCEL: Fichier '{EXCEL_FILE_NAME}' existe déjà.")

    except PermissionError:
        Logger.error(f"EXCEL: Erreur de permission. Impossible d'écrire dans '{excel_path}'. Vérifiez si le fichier est ouvert.")
    except Exception as e:
        Logger.error(f"EXCEL: Erreur lors de l'initialisation du fichier Excel: {e}")


def add_record_to_excel(data, action="UNKNOWN"):
    """Ajoute un enregistrement (ligne) au fichier Excel."""
    excel_path = get_excel_path()
    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook[SHEET_NAME] # Accéder à la feuille par son nom

        # Préparer la ligne de données dans le bon ordre
        row_data = [
            data.get('id', 'N/A'), # ID de la base de données si disponible
            action, # Type d'action (ADD, MOVE, etc.)
            data.get('timestamp', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            data.get('palette_number', 'N/A'),
            data.get('product_name', 'N/A'),
            data.get('price', 'N/A'),
            data.get('expiry_date', 'N/A'),
            data.get('lot_number', 'N/A'),
            data.get('boxes_per_package', 'N/A'),
            data.get('location_id', 'N/A')
        ]
        sheet.append(row_data)
        workbook.save(excel_path)
        Logger.info(f"EXCEL: Enregistrement ajouté pour palette {data.get('palette_number', '?')} (Action: {action}).")
        return True

    except FileNotFoundError:
        Logger.error(f"EXCEL: Fichier '{excel_path}' non trouvé. Tentative de réinitialisation...")
        init_excel() # Tenter de recréer le fichier
        # Essayer d'ajouter à nouveau après réinitialisation
        try:
            workbook = openpyxl.load_workbook(excel_path)
            sheet = workbook[SHEET_NAME]
            row_data = [
                data.get('id', 'N/A'), action,
                data.get('timestamp', datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                data.get('palette_number', 'N/A'), data.get('product_name', 'N/A'),
                data.get('price', 'N/A'), data.get('expiry_date', 'N/A'),
                data.get('lot_number', 'N/A'), data.get('boxes_per_package', 'N/A'),
                data.get('location_id', 'N/A')
            ]
            sheet.append(row_data)
            workbook.save(excel_path)
            Logger.info(f"EXCEL: Enregistrement ajouté après réinitialisation pour palette {data.get('palette_number', '?')}.")
            return True
        except Exception as e_retry:
            Logger.error(f"EXCEL: Échec de l'ajout même après réinitialisation: {e_retry}")
            return False
    except KeyError:
         Logger.error(f"EXCEL: La feuille '{SHEET_NAME}' n'existe pas dans '{excel_path}'. Vérifiez le fichier.")
         return False
    except PermissionError:
        Logger.error(f"EXCEL: Erreur de permission. Impossible d'écrire dans '{excel_path}'. Vérifiez si le fichier est ouvert.")
        return False
    except Exception as e:
        Logger.error(f"EXCEL: Erreur lors de l'ajout de l'enregistrement: {e}")
        return False

# init_excel() # Appeler depuis main.py
